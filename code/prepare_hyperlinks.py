import pandas as pd
import numpy as np
import re
import openpyxl

# get file path and read in a datframe of our data and the workbook form of our data
file_path = "data/nsf_grfp_examples.xlsx"
workbook = openpyxl.load_workbook(file_path) 
nsf_df = df = pd.read_excel(file_path)

# extract the sheet that contans hyperlinks to the data we want
ws = workbook['nsfgrfp']

#extract the hyperlinks 
proposal_hl = []
personal_hl = []
for i in range(2, len(nsf_df) + 2):
    try: 
        proposal_hl.append(ws.cell(row=i, column=6).hyperlink.target)
    except:
        if "=HYPERLINK" in ws.cell(row=i, column=6).value.split(",")[0]:
            proposal_hl.append(re.findall('"([^"]*)"', ws.cell(row=i, column=6).value.split(",")[0])[0])
        else:
            proposal_hl.append("NA")
    try: 
        personal_hl.append(ws.cell(row=i, column=7).hyperlink.target)
    except:
        if "=HYPERLINK" in ws.cell(row=i, column=7).value.split(",")[0]:
            personal_hl.append(re.findall('"([^"]*)"', ws.cell(row=i, column=7).value.split(",")[0])[0])
        else:
            personal_hl.append("not_here")   

#add them to the dataframe 
nsf_df = nsf_df.assign(proposal_hyperlinks=proposal_hl) 
nsf_df = nsf_df.assign(personal_hyperlinks=personal_hl)

# get rid of rows where there is no proposal available
nsf_df = nsf_df.replace("NA", np.NaN)
nsf_df = nsf_df[nsf_df['proposal_hyperlinks'].notna()]

nsf_df['proposal_id'] = nsf_df['proposal_hyperlinks'].str.split('/').str[5]
nsf_df['personal_id'] = nsf_df['personal_hyperlinks'].str.split('/').str[5]

csv_filename = 'data/hyperlinks.csv'
nsf_df.to_csv(csv_filename, index=False)
