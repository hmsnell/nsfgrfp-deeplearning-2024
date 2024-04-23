import pandas as pd
import openpyxl

# get file path and read in a datframe of our data and the workbook form of our data
file_path = "../data/nsf_grfp_examples.xlsx"
workbook = openpyxl.load_workbook(file_path) 
nsf_df = df = pd.read_excel(file_path)

# extract the sheet that contans hyperlinks to the data we want
ws = workbook['nsfgrfp']

#extract the hyperlinks 
proposal_hl = []
personal_hl = []
for i in range(2, len(nsf_df) + 2):
    try: 
        proposal_hl.append(ws.cell(row=i, column=6).hyperlink.target)
    except:
        proposal_hl.append("NA")
        
    try: 
        personal_hl.append(ws.cell(row=i, column=7).hyperlink.target)
    except:
        personal_hl.append("NA")   

#add them to the dataframe 
nsf_df = nsf_df.assign(proposal_hyperlinks=proposal_hl) 
nsf_df = nsf_df.assign(personal_hyperlinks=personal_hl) 
